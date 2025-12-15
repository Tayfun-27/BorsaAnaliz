import streamlit as st
import pandas as pd
import numpy as np
import io
import time
import os
import yfinance as yf
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from sklearn.linear_model import LinearRegression
from scipy.stats import pearsonr

# --- STREAMLIT SAYFA AYARLARI ---
st.set_page_config(page_title="Ultra Full Borsa Analiz", layout="wide")
st.title("🚀 Ultra Borsa Analiz: Tam Kapsamlı Versiyon")
st.markdown("Bu uygulama, Klasik Analiz ve Expert MA analizini hatasız veri temizliği ile birleştirir.")

# ==================== 1. HİSSE LİSTESİ ====================
def get_hardcoded_tickers():
    return [
        "ACSEL", "ADEL", "AEFES", "AFYON", "AGESA", "AGHOL", "AGYO", "AKBNK", "AKCNS", "AKENR", 
        "AKFGY", "AKGRT", "AKSA", "AKSEN", "AKSGY", "ALARK", "ALBRK", "ALCAR", "ALCTL", "ALFAS", 
        "ALGYO", "ALKIM", "ALMAD", "ANACM", "ANELE", "ANHYT", "ANSGR", "ARASE", "ARCLK", "ARDYZ", 
        "ARENA", "ARSAN", "ASELS", "ASTOR", "ASUZU", "ATAKP", "ATATP", "AVGYO", "AVOD", "AYDEM", 
        "AYEN", "AYGAZ", "BAGFS", "BAKAB", "BANVT", "BARMA", "BASGZ", "BERA", "BEYAZ", "BFREN", 
        "BIMAS", "BIOEN", "BIZIM", "BJKAS", "BOBET", "BOSSA", "BRISA", "BRKO", "BRSAN", "BRYAT", 
        "BSOKE", "BTCIM", "BUCIM", "BURCE", "BURVA", "BVSAN", "BYDNR", "CANTE", "CCOLA", "CEMAS", 
        "CEMTS", "CIMSA", "CLEBI", "CMBTN", "CMENT", "CONSE", "CRFSA", "CUSAN", "CVKMD", "CWENE", 
        "DAGI", "DAPGM", "DARDL", "DENGE", "DERIM", "DESA", "DEVA", "DGATE", "DGGYO", "DGNMO", 
        "DITAS", "DMSAS", "DNISI", "DOAS", "DOCO", "DOGUB", "DOHOL", "DOKTA", "DURDO", "DYOBY", 
        "EBEBK", "ECILC", "ECZYT", "EDATA", "EDIP", "EGEEN", "EGGUB", "EGPRO", "EGSER", "EKGYO", 
        "EKSUN", "ELITE", "EMKEL", "ENJSA", "ENKAI", "ENSRI", "EPLAS", "ERBOS", "EREGL", "ESCAR", 
        "ESCOM", "ESEN", "ETILR", "EUHOL", "EUPWR", "EUREN", "FADE", "FENER", "FLAP", "FMIZP", 
        "FONET", "FORMT", "FRIGO", "FROTO", "GARAN", "GEDIK", "GENIL", "GENTS", "GEREL", "GESAN", 
        "GLBMD", "GLRYH", "GLYHO", "GOKNR", "GOLTS", "GOODY", "GOZDE", "GSDHO", "GSRAY", "GUBRF", 
        "GWIND", "HALKB", "HATEK", "HDFGS", "HEDEF", "HEKTS", "HKTM", "HLGYO", "HTTBT", "HUNER", 
        "HURGZ", "ICBCT", "IEYHO", "IHAAS", "IHEVA", "IHGZT", "IHLAS", "IHLGM", "IHYAY", "IMASM", 
        "INDES", "INFO", "INGRM", "INTEM", "INVEO", "INVES", "IPEKE", "ISCTR", "ISDMR", "ISFIN", 
        "ISGSY", "ISGYO", "ISKPL", "ISMEN", "ISSEN", "ISYAT", "ITTFH", "IZENR", "IZFAS", "IZMDC", 
        "JANTS", "KAPLM", "KAREL", "KARSN", "KARTN", "KARYE", "KATMR", "KAYSE", "KCAER", "KCHOL", 
        "KENT", "KERVN", "KERVT", "KFEIN", "KGYO", "KIMMR", "KLGYO", "KLKIM", "KLMSN", "KLRHO", 
        "KLSER", "KMPUR", "KNFRT", "KONKA", "KONTR", "KONYA", "KOPOL", "KORDS", "KOZAA", "KOZAL", 
        "KRDMA", "KRDMB", "KRDMD", "KRGYO", "KRONT", "KRPLS", "KRSTL", "KRTEK", "KSTUR", "KTLEV", 
        "KUTPO", "KUYAS", "KZBGY", "LIDER", "LIDFA", "LINK", "LKMNH", "LOGO", "LUKSK", "MAALT", 
        "MACKO", "MAGEN", "MAKIM", "MAKTK", "MANAS", "MARKA", "MARTI", "MAVI", "MEDTR", "MEGAP", 
        "MEPET", "MERCN", "MERKO", "METRO", "MGROS", "MIATK", "MIPAZ", "MNDRS", "MOBTL", "MPARK", 
        "MRGYO", "MRSHL", "MSGYO", "MTRKS", "MTRYO", "MZHLD", "NATEN", "NETAS", "NIBAS", "NTGAZ", 
        "NTHOL", "NUGYO", "NUHCM", "OBASE", "ODAS", "ONCSM", "ORCAY", "ORGE", "OSMEN", "OSTIM", 
        "OTKAR", "OYAKC", "OYAYO", "OYLUM", "OZGYO", "OZKGY", "OZRDN", "OZSUB", "PAGYO", "PAMEL", 
        "PAPIL", "PARSN", "PASEU", "PCILT", "PEGYO", "PEKGY", "PENGD", "PETKM", "PETUN", "PGSUS", 
        "PINSU", "PKART", "PKENT", "PLTUR", "PNLSN", "PNSUT", "POLHO", "POLTK", "PRDGS", "PRKAB", 
        "PRKME", "PSDTC", "PSGYO", "QNBFB", "QUAGR", "RALYH", "RAYSG", "RNPOL", "RTALB", "RUBNS", 
        "RYGYO", "RYSAS", "SAHOL", "SAMAT", "SANEL", "SANKO", "SARKY", "SASA", "SAYAS", "SDTTR", 
        "SEKFK", "SEKUR", "SELEC", "SELGD", "SELVA", "SEYKM", "SILVR", "SISE", "SKBNK", "SKTAS", 
        "SMART", "SMRTG", "SNGYO", "SNKRN", "SNPAM", "SOKE", "SOKM", "SONME", "SRVGY", "SUMAS", 
        "SUNGW", "SUVEN", "TATGD", "TAVHL", "TBORG", "TCELL", "TDGYO", "TEKTU", "TERA", "TEZOL", 
        "THYAO", "TKFEN", "TKNSA", "TLMAN", "TMPOL", "TMSN", "TNZTP", "TOASO", "TRCAS", "TRGYO", 
        "TRILC", "TSGYO", "TSKB", "TSPOR", "TTKOM", "TTRAK", "TUCLK", "TUKAS", "TUPRS", "TURGG", 
        "TURSG", "ULAS", "ULKER", "ULUFA", "ULUSE", "ULUUN", "UNLU", "USAK", "VAKBN", "VAKFN", 
        "VAKKO", "VANGD", "VBTYZ", "VERTU", "VERUS", "VESBE", "VESTL", "VKFYO", "VKGYO", "VKING", 
        "YAPRK", "YATAS", "YAYLA", "YEOTK", "YESIL", "YGGYO", "YGYO", "YKBNK", "YKSLN", "YONGA", 
        "YUNSA", "YYAPI", "ZEDUR", "ZOREN"
    ]

# ==================== 2. VERİ YÜKLEME VE TEMİZLEME ====================
def download_data_task(status_box, progress_bar):
    output_folder = "Borsa_Verileri"
    if not os.path.exists(output_folder): os.makedirs(output_folder)
    
    tickers = get_hardcoded_tickers()
    total = len(tickers)
    success = 0
    errors = []
    
    # Session oluşturma kısmını SİLDİK. yfinance kendi halledecek.

    status_box.info("İndirme işlemi başlatılıyor (YFinance Auto-Session)...")
    
    for i, sym in enumerate(tickers):
        try:
            status_box.text(f"İndiriliyor: {sym} ({i+1}/{total})")
            progress_bar.progress((i+1)/total)
            
            yf_symbol = f"{sym}.IS"
            
            # session parametresini KALDIRDIK
            ticker_obj = yf.Ticker(yf_symbol) 
            df = ticker_obj.history(period="5y")
            
            if df.empty:
                # session parametresini KALDIRDIK
                df = yf.download(yf_symbol, period="5y", progress=False)
                if df.empty:
                    continue

            if len(df) < 233: 
                continue
            
            # Format düzenleme
            df = df.reset_index()
            if 'Date' not in df.columns and 'Datetime' in df.columns:
                df = df.rename(columns={'Datetime': 'Date'})
            
            df['Date'] = pd.to_datetime(df['Date']).dt.date
            
            df = df.rename(columns={
                "Date": "DATE", "Close": "CLOSING_TL", "Open": "OPEN_TL", 
                "High": "HIGH_TL", "Low": "LOW_TL", "Volume": "VOLUME"
            })
            
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(0)

            cols = ["DATE","CLOSING_TL","OPEN_TL","HIGH_TL","LOW_TL","VOLUME"]
            available_cols = [c for c in cols if c in df.columns]
            df = df[available_cols]
            
            df = df.dropna(subset=['CLOSING_TL'])
            
            if df.empty: continue

            df.to_excel(os.path.join(output_folder, f"{sym}.xlsx"), index=False)
            success += 1
            
            time.sleep(0.1)
            
        except Exception as e:
            errors.append(f"{sym} Hatası: {str(e)}")
            time.sleep(0.1)
            pass
            
    if success == 0 and errors:
        st.error("Hiçbir hisse indirilemedi! Alınan örnek hatalar:")
        for err in errors[:3]:
            st.write(err)
            
    return output_folder, success

def load_stock_df(path):
    try:
        df = pd.read_excel(path, engine="openpyxl")
        df.columns = [str(c).strip().upper() for c in df.columns]
        if "DATE" in df.columns:
            df["DATE"] = pd.to_datetime(df["DATE"], errors="coerce", dayfirst=True)
            df = df.dropna(subset=["DATE"])
            df = df.sort_values("DATE")
        
        # EKSTRA GÜVENLİK: Kapanış fiyatı boşsa doldur veya sil
        if "CLOSING_TL" in df.columns:
            df = df.dropna(subset=["CLOSING_TL"]) # Boşları sil
            # df["CLOSING_TL"] = df["CLOSING_TL"].ffill() # Alternatif: Önceki değerle doldur
            
        return df
    except: return None

def autofit(ws, df):
    if df is None or df.empty: return
    for i, c in enumerate(df.columns, 1):
        try:
            l = max(df[c].astype(str).apply(len).max(), len(str(c)))
            ws.column_dimensions[get_column_letter(i)].width = min(l+2, 50)
        except: pass

# ==================== 3. HESAPLAMA MOTORU ====================
def calc_ema(d, p): return d.ewm(span=p, adjust=False).mean()

# Expert Math
def linreg(p, l, o):
    if len(p)<l: return np.nan
    x=np.arange(l); y=p[-l:]; 
    # NaN kontrolü
    if np.isnan(y).any(): return np.nan
    
    s, i = np.polyfit(x, y, 1)
    return s*(l-1+o)+i

def zlsma(p,l,o):
    ps=pd.Series(p); ls=ps.rolling(l).apply(lambda x:linreg(x,l,0),raw=False)
    return (ls+(ls-ls.rolling(l).apply(lambda x:linreg(x,l,0),raw=False))).shift(-o)
def smma(d,l):
    s=np.zeros(len(d)); s[0]=d.iloc[0]
    for i in range(1,len(d)): s[i]=(s[i-1]*(l-1)+d.iloc[i])/l
    return s
def ma_type(d,p): e1=calc_ema(d,p); e2=calc_ema(e1,p); e3=calc_ema(e2,p); return 3*e1-3*e2+e3
def m1(d,a): e1=d.ewm(alpha=a,adjust=False).mean(); e2=e1.ewm(alpha=a,adjust=False).mean(); return 2*e1-e2
def lr_slope(d,l):
    if len(d)<l: return np.nan
    y=d[-l:]; x=np.arange(l)
    if np.isnan(y).any(): return np.nan # Hata önleyici
    return np.polyfit(x,y,1)[0]
def perc(d,w,p): return d.rolling(w).quantile(p/100, interpolation='nearest')
def finh(d,p): h=p//2; return calc_ema(2*calc_ema(d,h)-calc_ema(d,p), int(np.sqrt(p)))
def hma(d,p):
    def w(x,l): wt=np.arange(1,l+1); return np.dot(x,wt)/wt.sum()
    h=round(p/2); w1=d.rolling(h).apply(lambda x:w(x,h),raw=True); w2=d.rolling(p).apply(lambda x:w(x,p),raw=True)
    return (2*w1-w2).rolling(round(np.sqrt(p))).apply(lambda x:w(x,round(np.sqrt(p))),raw=True)
def jma(d,l=50,ph=1,po=4):
    pr=ph/100+1.5 if ph<=100 else 2.5; b=0.45*(l-1)/(0.45*(l-1)+2); a=b**po
    j=np.zeros(len(d)); e0=np.zeros(len(d)); e1=np.zeros(len(d)); e2=np.zeros(len(d))
    for i in range(len(d)):
        c=d.iloc[i]
        if np.isnan(c): c = j[i-1] if i>0 else 0 # NaN koruması
        if i==0: e0[i]=c; e2[i]=c; j[i]=c
        else:
            e0[i]=(1-a)*c+a*e0[i-1]; e1[i]=(c-e0[i])*(1-b)+b*e1[i-1]
            e2[i]=(e0[i]+pr*e1[i]-j[i-1])*((1-a)**2)+((a**2)*e2[i-1]); j[i]=e2[i]+j[i-1]
    return j
def tema(d,l): e1=calc_ema(d,l); e2=calc_ema(e1,l); return 3*(e1-e2)+calc_ema(e2,l)
def dema(d,l): e1=calc_ema(d,l); return 2*e1-calc_ema(e1,l)
def ama(d,p):
    a=np.zeros(len(d)); f=2/6; s=2/16
    for i in range(len(d)):
        if i<p: a[i]=d.iloc[i]; continue
        dS=abs(d.iloc[i]-d.iloc[i-p]); dN=sum(abs(d.iloc[j]-d.iloc[j-1]) for j in range(i-p+1,i+1))
        ER=dS/dN if dN!=0 else 0; ssc=ER*(f-s)+s
        a[i]=(ssc**2)*(d.iloc[i]-a[i-1])+a[i-1]
    return a
def calc_macd(d,f,s,sm): m=calc_ema(d,f)-calc_ema(d,s); sig=calc_ema(m,sm); return m, sig

# ==================== 4. RAPORLAMA YARDIMCILARI ====================
def write_comp_table(ws, df, title, row):
    if df.empty: ws.cell(row,1,f"--- {title} (Veri Yok) ---").font=Font(bold=True); return row+2
    ws.cell(row,1,f"--- {title} ---").font=Font(bold=True, size=12); row+=1
    h_fill=PatternFill("solid",fgColor="BFBFBF")
    for i,c in enumerate(df.columns,1): ws.cell(row,i,c).font=Font(bold=True); ws.cell(row,i).fill=h_fill
    for r_idx, r_val in enumerate(df.values, start=row+1):
        fill=PatternFill("solid",fgColor="D9E1F2" if (r_idx-row)%2==0 else "E2EFDA")
        for i, val in enumerate(r_val,1):
            c=ws.cell(r_idx,i,val); c.fill=fill
            if "Yeni" in str(val): c.font=Font(color="008000", bold=True)
            elif "Çıktı" in str(val): c.font=Font(color="FF0000", bold=True)
            elif "Statü Değişti" in str(val): c.font=Font(color="FF8C00", bold=True)
    autofit(ws, df); return ws.max_row+2

def write_expert_table(ws, title, data, headers, b_fill, b_font, cond=False):
    if not data: return
    sr = ws.max_row + 2 if ws.max_row > 1 else 1
    cell = ws.cell(sr, 1, title); cell.font = b_font; cell.fill = b_fill; cell.alignment = Alignment('center', 'center')
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(headers))
    rh = sr + 1
    for i, h in enumerate(headers, 1):
        c = ws.cell(rh, i, h); c.fill = b_fill; c.font = b_font; c.alignment = Alignment('center', 'center')
    rd = rh + 1
    o_fill = PatternFill("solid", fgColor="FFA500"); bl_font = Font(color="000000", bold=True)
    for row in data:
        cc = row.get('Son Kapanış')
        for i, h in enumerate(headers, 1):
            val = row.get(h)
            c = ws.cell(rd, i, val); c.fill = b_fill; c.font = b_font
            if cond and row.get('Puan') in [12, 13]:
                met = True
                if h in ['ZLSMA','SMMA','MA1','MA2','M1','Percentile','FINH','HMA','JMA','TEMA','DEMA','AMA']:
                    if not pd.isna(val) and isinstance(cc,(int,float)) and not (cc > float(val)): met = False
                elif h in ['LinReg','MACD Kesişme']:
                    if val != "Pozitif": met = False
                if not met and h not in ['Hisse','Son Kapanış','Puan','Önceki Puan']:
                    c.fill = o_fill; c.font = bl_font
        rd += 1
    ws.append([]); ws.append([])

# ==================== 5. SEKMELER (TABS) ====================

tab_veri, tab_klasik, tab_expert = st.tabs(["1. Veri İndir", "2. Klasik Analiz (Hızlı)", "3. Expert Analiz (Detaylı)"])

# --- TAB 1: VERİ İNDİRME ---
with tab_veri:
    st.header("Veri İndirme Merkezi")
    st.info("Analiz yapabilmek için önce güncel verileri indirmeniz gerekir.")
    
    if st.button("Verileri İndir / Güncelle", type="primary"):
        status = st.empty()
        prog = st.progress(0)
        fld, cnt = download_data_task(status, prog)
        status.success(f"İşlem Tamam! {cnt} adet hisse 'Borsa_Verileri' klasörüne indirildi.")
        st.write("Şimdi diğer sekmelere geçerek analiz yapabilirsiniz.")

# --- TAB 2: KLASİK ANALİZ ---
with tab_klasik:
    st.header("Klasik Analiz (EMA & Kanal)")
    st.markdown("Trend, EMA dizilimleri ve Kanal takibi yapar. Daha hızlı çalışır.")
    
    prev_rep = st.file_uploader("Önceki Klasik Rapor (İsteğe Bağlı - Kıyaslama İçin)", type=['xlsx'], key="cls_up")
    
    if st.button("Klasik Analizi Başlat"):
        fld = "Borsa_Verileri"
        if not os.path.exists(fld) or not os.listdir(fld):
            st.error("Önce 'Veri İndir' sekmesinden veri indirmelisiniz!")
        else:
            st.info("Klasik analiz yapılıyor...")
            start = time.time()
            files = [f for f in os.listdir(fld) if f.endswith('.xlsx')]
            
            res_ema = []; ch_res = []; pea_res = {}
            prog = st.progress(0); tot = len(files)
            
            for i, f in enumerate(files):
                prog.progress((i+1)/tot)
                df = load_stock_df(os.path.join(fld, f))
                if df is None or len(df) < 233: continue
                nm = f.replace('.xlsx', ''); cl = df['CLOSING_TL']
                
                # EMA
                for p in [8,13,21,34,55,89,144,233]: df[f'EMA{p}'] = calc_ema(cl, p)
                last = df.iloc[-1]; cp = last['CLOSING_TL']
                up = all(cp > last[f'EMA{p}'] for p in [8,13,21,34,55,89,144,233])
                id_up = (cp > last['EMA8'] > last['EMA13'] > last['EMA21'] > last['EMA34'] > last['EMA55'] > last['EMA89'] > last['EMA144'] > last['EMA233'])
                
                rd = {'Stock Name': nm, 'Closing Price': cp, 'Status': 'UP' if up else '', 'Ideal Status': 'IDEAL UP' if id_up else ''}
                for p in [8,13,21,34,55,89,144,233]: rd[f'EMA{p}'] = last[f'EMA{p}']
                res_ema.append(rd)
                
                # Pearson
                p_dict = {}
                for p in [55,144,233,377,610,987]:
                    if len(df) >= p:
                        y = cl.tail(p); x = np.arange(p).reshape(-1, 1)
                        # NaNs kontrolü
                        if y.isnull().any():
                             p_dict[f'{p} Gün'] = np.nan
                        else:
                             p_dict[f'{p} Gün'] = np.corrcoef(x.T[0], y)[0,1] if len(y)>2 else np.nan
                    else: p_dict[f'{p} Gün'] = np.nan
                pea_res[nm] = p_dict
                
                # Kanal (HATA ÇÖZÜMÜ BURADA: NaN değerler temizleniyor)
                for v in [144,233,377,610]:
                    if len(df) >= v:
                        vdf = df.tail(v)
                        # Temizlik
                        vdf = vdf.dropna(subset=['CLOSING_TL'])
                        if len(vdf) < 5: continue # Yeterli veri yoksa atla
                        
                        y = vdf['CLOSING_TL'].values
                        x = np.arange(len(y)).reshape(-1, 1)
                        
                        try:
                            md = LinearRegression().fit(x, y)
                            pr = md.predict(x); std = np.std(y - pr)
                            u = pr[-1] + 2*std; l = pr[-1] - 2*std
                            pc = pearsonr(vdf['CLOSING_TL'], pr)[0] if len(y)>2 else np.nan
                            if md.coef_[0] < 0 and not np.isnan(pc): pc = -pc
                            ch_res.append({'Hisse': nm, 'Vade': v, 'Son': cp, 'Ust': u, 'Alt': l, 'Ust%': (u-cp)/cp*100, 'Alt%': (cp-l)/cp*100, 'Pearson': pc})
                        except: pass # Hata verirse (örn. tüm değerler aynıysa) atla
            
            # Excel Yazma (Klasik)
            bio = io.BytesIO()
            with pd.ExcelWriter(bio, engine='openpyxl') as w:
                # 1. EMA
                df_e = pd.DataFrame(res_ema); df_e.to_excel(w, sheet_name='EMA_Sonuclari', index=False)
                ws = w.sheets['EMA_Sonuclari']; c_b = PatternFill("solid", fgColor="0000FF"); c_o = PatternFill("solid", fgColor="FFA500"); fw = Font(color="FFFFFF", bold=True)
                for r in range(2, ws.max_row+1):
                    if df_e.iloc[r-2]['Status'] == 'UP': 
                        for c in range(1, ws.max_column+1): ws.cell(r,c).fill = c_b; ws.cell(r,c).font = fw
                    if df_e.iloc[r-2]['Ideal Status'] == 'IDEAL UP': 
                        c = ws.cell(r, df_e.columns.get_loc('Ideal Status')+1); c.fill = c_o; c.font = fw
                autofit(ws, df_e)
                
                # 2. UP-Ideal UP
                ws_u = w.book.create_sheet("UP-Ideal UP"); r = 1
                id_df = df_e[df_e['Ideal Status'] == 'IDEAL UP'].drop(columns=['Status', 'Ideal Status'])
                up_df = df_e[(df_e['Status'] == 'UP') & (df_e['Ideal Status'] != 'IDEAL UP')].drop(columns=['Status', 'Ideal Status'])
                if not id_df.empty:
                    ws_u.cell(r,1,"--- IDEAL UP ---").font = Font(bold=True); r+=1; ws_u.append(list(id_df.columns)); r+=1
                    for row in id_df.values: ws_u.append(list(row)); r+=1
                r+=2
                if not up_df.empty:
                    ws_u.cell(r,1,"--- UP ---").font = Font(bold=True); r+=1; ws_u.append(list(up_df.columns)); r+=1
                    for row in up_df.values: ws_u.append(list(row)); r+=1
                autofit(ws_u, df_e)

                # 3. Pearson
                df_p = pd.DataFrame(pea_res).T; df_p.to_excel(w, sheet_name='Pearson_Sonuclari'); autofit(w.sheets['Pearson_Sonuclari'], df_p)
                
                # 4. En Yuksek Pearson
                best = [{'Hisse':s, 'Periyot':max(v, key=v.get), 'Pearson':max(v.values())} for s,v in pea_res.items() if any(not pd.isna(x) for x in v.values())]
                if best: df_b = pd.DataFrame(best); df_b.to_excel(w, sheet_name='En_Yuksek_Pearson', index=False); autofit(w.sheets['En_Yuksek_Pearson'], df_b)
                
                # 5. Pozitif Pearson
                pos = {k:v for k,v in pea_res.items() if all(not pd.isna(x) and x > 0 for x in v.values())}
                if pos: 
                    df_pos = pd.DataFrame(pos).T; df_pos.to_excel(w, sheet_name='Pozitif_Pearson')
                    ws_p = w.sheets['Pozitif_Pearson']; nav = PatternFill("solid", fgColor="000080")
                    for r in range(2, ws_p.max_row+1):
                        for c in range(2, ws_p.max_column+1):
                            if isinstance(ws_p.cell(r,c).value, (int, float)) and ws_p.cell(r,c).value >= 0.9: ws_p.cell(r,c).fill = nav; ws_p.cell(r,c).font = fw
                    autofit(ws_p, df_pos)

                # 6. Com144.. & IdealUp Sheetleri
                ideal_dfs = {}
                ideal_stocks = set(df_e[df_e['Ideal Status']=='IDEAL UP']['Stock Name'])
                for p in [144, 233, 377, 610, 987]:
                    rows = []
                    for s in ideal_stocks:
                        if s in pea_res and pea_res[s].get(f'{p} Gün', 0) > 0.85:
                            cp = df_e[df_e['Stock Name']==s].iloc[0]['Closing Price']
                            rows.append({'Hisse':s, 'Kapanış':cp, 'Pearson':pea_res[s][f'{p} Gün']})
                    ideal_dfs[p] = pd.DataFrame(rows)
                
                s1 = set(ideal_dfs[144]['Hisse']) if not ideal_dfs[144].empty else set()
                s2 = set(ideal_dfs[233]['Hisse']) if not ideal_dfs[233].empty else set()
                s3 = set(ideal_dfs[377]['Hisse']) if not ideal_dfs[377].empty else set()
                common = s1.intersection(s2).intersection(s3)
                if common:
                    c_data = [{'Hisse':s, 'Kapanış':df_e[df_e['Stock Name']==s].iloc[0]['Closing Price'], 'Pearson 144':pea_res[s].get('144 Gün'), 'Pearson 233':pea_res[s].get('233 Gün'), 'Pearson 377':pea_res[s].get('377 Gün')} for s in common]
                    df_com = pd.DataFrame(c_data); df_com.to_excel(w, sheet_name='Com144-233-377', index=False); autofit(w.sheets['Com144-233-377'], df_com)
                else: w.book.create_sheet("Com144-233-377").cell(1,1,"Veri Yok")
                
                for p in [144, 233, 377, 610, 987]:
                    sn = f"{p}IdealUp"
                    if not ideal_dfs[p].empty: ideal_dfs[p].to_excel(w, sheet_name=sn, index=False); autofit(w.sheets[sn], ideal_dfs[p])
                    else: w.book.create_sheet(sn).cell(1,1,"Veri Yok")

                # 7. Rapor_Upd
                ws_r = w.book.create_sheet("Rapor_Upd"); r_idx = 1
                if prev_rep:
                    try:
                        old_ema = pd.read_excel(prev_rep, sheet_name='EMA_Sonuclari', engine='openpyxl')
                        n_id = set(df_e[df_e['Ideal Status']=='IDEAL UP']['Stock Name'])
                        o_id = set(old_ema[old_ema['Ideal Status']=='IDEAL UP']['Stock Name']) if 'Ideal Status' in old_ema else set()
                        chg = []
                        for s in n_id.union(o_id):
                            stt = "Yeni (Ideal UP)" if s in n_id and s not in o_id else "Çıktı (Ideal UP)" if s in o_id and s not in n_id else "Değişmedi"
                            if stt != "Değişmedi": chg.append({'Hisse':s, 'Durum':stt, 'Son Fiyat': df_e[df_e['Stock Name']==s].iloc[0]['Closing Price'] if s in n_id else 0})
                        write_comp_table(ws_r, pd.DataFrame(chg), "Ideal UP Değişimleri", 1)
                    except: ws_r.cell(1,1,"Önceki rapor okunamadı")
                else: ws_r.cell(1,1,"Kıyaslama dosyası yok")

                # 8. Kanal & ListeBaşı
                if ch_res:
                    df_c = pd.DataFrame(ch_res); df_c['Dikkat'] = ''; df_c.to_excel(w, sheet_name='Kanal_Ekstra', index=False)
                    ws = w.sheets['Kanal_Ekstra']
                    for r in range(2, ws.max_row+1):
                        rd = df_c.iloc[r-2]
                        if rd['Pearson'] > 0.9:
                            for c in range(1, ws.max_column+1): ws.cell(r,c).fill=PatternFill("solid",fgColor="000080"); ws.cell(r,c).font=fw
                            if rd['Ust%']<=2 or rd['Alt%']<=2: ws.cell(r, ws.max_column).fill=PatternFill("solid",fgColor="FF0000"); ws.cell(r, ws.max_column).font=fw
                    autofit(ws, df_c)
                    
                    df_lb = df_c[(df_c['Pearson']>0.9) & ((df_c['Ust%']<=2)|(df_c['Alt%']<=2))]
                    if not df_lb.empty: df_lb.to_excel(w, sheet_name='ListeBaşı', index=False); autofit(w.sheets['ListeBaşı'], df_lb)

            st.success("Klasik Analiz Tamamlandı!")
            st.download_button("📥 Klasik Raporu İndir", bio.getvalue(), f"Klasik_Analiz_{datetime.now().strftime('%Y%m%d')}.xlsx")

# --- TAB 3: EXPERT ANALİZ ---
with tab_expert:
    st.header("Expert Analiz (Puanlama & Detay)")
    st.markdown("14 farklı indikatörle puanlama yapar. **Çok işlem yaptığı için uzun sürer.**")
    
    prev_exp = st.file_uploader("Önceki Expert Raporu (İsteğe Bağlı - Kıyaslama İçin)", type=['xlsx'], key="exp_up")
    
    if st.button("Expert Analizi Başlat"):
        fld = "Borsa_Verileri"
        if not os.path.exists(fld) or not os.listdir(fld):
            st.error("Önce 'Veri İndir' sekmesinden veri indirmelisiniz!")
        else:
            st.info("Expert analiz yapılıyor... (Bu işlem uzun sürebilir)")
            files = [f for f in os.listdir(fld) if f.endswith('.xlsx')]
            res = []; prog = st.progress(0); tot = len(files)
            
            for i, f in enumerate(files):
                prog.progress((i+1)/tot)
                df = load_stock_df(os.path.join(fld, f))
                if df is None or len(df) < 200: continue
                df.set_index('DATE', inplace=True); cl = df['CLOSING_TL']
                
                check = {
                    'ZLSMA': zlsma(cl,173,0).iloc[-1], 'SMMA': smma(cl,120)[-1], 'MA1': ma_type(cl,107).iloc[-1],
                    'MA2': ma_type(cl,120).iloc[-1], 'M1': m1(cl,0.023).iloc[-1], 'Percentile': perc(cl,44,89).iloc[-1],
                    'FINH': finh(cl,89).iloc[-1], 'HMA': hma(cl,196).iloc[-1], 'JMA': jma(cl)[-1],
                    'TEMA': tema(cl,144).iloc[-1], 'DEMA': dema(cl,89).iloc[-1], 'AMA': ama(cl,5)[-1]
                }
                ls = "Pozitif" if lr_slope(cl,105)>0 else "Negatif"
                md, sig = calc_macd(cl,49,55,5)
                ms = "Pozitif" if md.iloc[-1]>sig.iloc[-1] else "Negatif"
                sc = sum(1 for v in check.values() if not pd.isna(v) and cl.iloc[-1]>v) + (1 if ls=="Pozitif" else 0) + (1 if ms=="Pozitif" else 0)
                res.append({'Hisse': f.replace('.xlsx',''), 'Son Kapanış': cl.iloc[-1], **check, 'LinReg': ls, 'MACD Kesişme': ms, 'Puan': sc})
            
            # Expert Excel Yazma
            bio = io.BytesIO()
            with pd.ExcelWriter(bio, engine='openpyxl') as w:
                p_map = {}
                if prev_exp:
                    try: 
                        odf = pd.read_excel(prev_exp, sheet_name='Expert_Dashboard')
                        p_map = odf.set_index('Hisse')['Puan'].fillna(-1).astype(int).to_dict()
                    except: pass
                
                r14=[x for x in res if x['Puan']==14]; r13=[x for x in res if x['Puan']==13]; r12=[x for x in res if x['Puan']==12]
                n12=[]; r1213=[]; r1314=[]; r1214=[]; d13=[]; d14=[]
                dr1413=[]; dr1412=[]; dr1312=[]; dr14b=[]; dr13b=[]; dr12b=[]
                
                for r in res:
                    h=r['Hisse']; p=r['Puan']; pp=p_map.get(h,-1); e={'Hisse':h,'Son Kapanış':r['Son Kapanış'],'Puan':p,'Önceki Puan':pp if pp!=-1 else 'Yok'}
                    if p==12 and pp<12: n12.append(e)
                    if p==13 and pp==12: r1213.append(e)
                    if p==14 and pp==13: r1314.append(e)
                    if p==14 and pp==12: r1214.append(e)
                    if p==13 and pp<12: d13.append(e)
                    if p==14 and pp<12: d14.append(e)
                    if pp>=12 and p>=12 and p<pp:
                        if pp==14 and p==13: dr1413.append(e)
                        if pp==14 and p==12: dr1412.append(e)
                        if pp==13 and p==12: dr1312.append(e)
                
                for h, pp in p_map.items():
                    curr = next((x for x in res if x['Hisse']==h), None)
                    if curr and curr['Puan']<12 and pp>=12:
                        e = {'Hisse':h, 'Son Kapanış':curr['Son Kapanış'], 'Puan':curr['Puan'], 'Önceki Puan': pp}
                        if pp==14: dr14b.append(e)
                        if pp==13: dr13b.append(e)
                        if pp==12: dr12b.append(e)

                wsd=w.book.create_sheet("Expert_Dashboard"); hd=list(res[0].keys()) if res else []
                wsd.append(hd)
                for r in res: wsd.append([r.get(k) for k in hd])
                hf=PatternFill("solid",fgColor="4F81BD"); wf=Font(color="FFFFFF",bold=True); nf=PatternFill("solid",fgColor="000080")
                for c in wsd[1]: c.fill=hf; c.font=wf
                for row in wsd.iter_rows(min_row=2):
                    try:
                        cv=float(row[1].value)
                        for i in [2,3,4,5,6,8,9,10,11,13,14,15]:
                            if row[i].value and cv>float(row[i].value): row[i].fill=nf; row[i].font=wf
                        if row[7].value=="Pozitif": row[7].fill=nf; row[7].font=wf
                        if row[12].value=="Pozitif": row[12].fill=nf; row[12].font=wf
                    except: pass
                autofit(wsd, pd.DataFrame(res))

                wsr=w.book.create_sheet("Expert_SONUÇ")
                pet=PatternFill("solid",fgColor="004F5C"); bro=PatternFill("solid",fgColor="A52A2A"); rol=PatternFill("solid",fgColor="0A4F32")
                blu=PatternFill("solid",fgColor="0000FF"); grn=PatternFill("solid",fgColor="008000"); red=PatternFill("solid",fgColor="FF0000")
                ch=['Hisse','Son Kapanış','Puan','Önceki Puan']
                
                write_expert_table(wsr, "14 Puan Alanlar", r14, hd, pet, wf)
                write_expert_table(wsr, "13 Puan Alanlar", r13, hd, bro, wf, True)
                write_expert_table(wsr, "12 Puan Alanlar", r12, hd, rol, wf, True)
                write_expert_table(wsr, "Yeni Girenler (12)", n12, ch, blu, wf)
                write_expert_table(wsr, "Yükselenler (12->13)", r1213, ch, grn, wf)
                write_expert_table(wsr, "Yükselenler (13->14)", r1314, ch, grn, wf)
                write_expert_table(wsr, "Yükselenler (12->14)", r1214, ch, grn, wf)
                write_expert_table(wsr, "Direkt Girenler (13)", d13, ch, blu, wf)
                write_expert_table(wsr, "Direkt Girenler (14)", d14, ch, blu, wf)
                write_expert_table(wsr, "Düşenler (14->13)", dr1413, ch, red, wf)
                write_expert_table(wsr, "Düşenler (14->12)", dr1412, ch, red, wf)
                write_expert_table(wsr, "Düşenler (13->12)", dr1312, ch, red, wf)
                write_expert_table(wsr, "Liste Dışı (14->xx)", dr14b, ch, red, wf)
                write_expert_table(wsr, "Liste Dışı (13->xx)", dr13b, ch, red, wf)
                write_expert_table(wsr, "Liste Dışı (12->xx)", dr12b, ch, red, wf)
                for i in range(1,20): wsr.column_dimensions[get_column_letter(i)].width=16
            
            st.success("Expert Analiz Tamamlandı!")
            st.download_button("📥 Expert Raporu İndir", bio.getvalue(), f"Expert_Analiz_{datetime.now().strftime('%Y%m%d')}.xlsx")